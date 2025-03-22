import openpyxl
from statistics import mean

def guardar_datos_estudiantes_en_excel():
    estudiantes = {}

    for i in range(3):
        nombre = input(f"Ingrese el nombre del estudiante {i+1}: ")
        notas = []
        for j in range(3):  
            nota = float(input(f"Ingrese la nota {j+1} de {nombre}: "))
            notas.append(nota)
        estudiantes[nombre] = notas

    libro = openpyxl.Workbook()
    hoja = libro.active

    fila = 1
    for nombre, notas in estudiantes.items():
        hoja.cell(row=fila, column=1, value=nombre)  
        hoja.cell(row=fila, column=2, value=mean(notas))  
        fila += 1

    libro.save("notas_estudiantes.xlsx")
    print("Los datos se han guardado en notas_estudiantes.xlsx")

if __name__ == "__main__":
    guardar_datos_estudiantes_en_excel()